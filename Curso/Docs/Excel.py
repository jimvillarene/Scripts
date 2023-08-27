#! /usr/bin/env python3
import openpyxl, os
dir_path = os.path.dirname(os.path.realpath(__file__))
excelFile=os.path.join(dir_path,'ExcelTest.xlsx')
workbook=openpyxl.load_workbook(excelFile)
print(type(workbook)) #Al importarse el wb se vuelve un pobjeto de tipo WOrkbook

print(workbook.get_sheet_names())
sheet=workbook.get_sheet_by_name('Sheet1')

cell=sheet['B2']
print(cell.value)

cell=sheet.cell(row=2,column=3)
print(cell.value)

for i in range(1,8):
    print(i,sheet.cell(row=i,column=2).value)

wb=openpyxl.Workbook()
print(wb.get_sheet_names())
sheet=wb.get_sheet_by_name('Sheet')
sheet['A1']=42
sheet['A2']='Hello'
saveFile=os.path.join(dir_path,'Example.xlsx')
wb.save(saveFile)
sheet2=wb.create_sheet()
print(wb.get_sheet_names())
sheet2.title='My new sheet name'
print(wb.get_sheet_names())